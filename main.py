from playwright.sync_api import sync_playwright
import pandas as pd
from datetime import date, timedelta
import requests
import os
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# Accommodation filter codes for Booking.com
ACCOMMODATION_FILTERS = {
    "hotel": "ht_id=204",
    "apartment": "ht_id=201",
    "resort": "ht_id=208",
    "hostel": "ht_id=203",
    "villa": "ht_id=213",
    "guesthouse": "ht_id=216",
}

def download_image(url, folder, filename):
    try:
        os.makedirs(folder, exist_ok=True)
        filepath = os.path.join(folder, filename)
        r = requests.get(url, timeout=10)
        if r.status_code == 200:
            with open(filepath, "wb") as f:
                f.write(r.content)
            return filepath
    except Exception as e:
        print(f"⚠ Error downloading image {url}: {e}")
    return None

def embed_images_in_excel(excel_file):
    """Replace image paths with actual embedded images inside Excel file."""
    wb = load_workbook(excel_file)
    ws = wb.active

    # Find Image Path column index
    img_col_idx = None
    for idx, col in enumerate(ws[1], start=1):
        if col.value == "Image Path":
            img_col_idx = idx
            break

    if img_col_idx:
        for row_idx in range(2, ws.max_row + 1):
            img_path = ws.cell(row=row_idx, column=img_col_idx).value
            if img_path and os.path.exists(img_path):
                try:
                    img = XLImage(img_path)
                    img.width = 80
                    img.height = 80
                    cell_address = f"{ws.cell(row=row_idx, column=img_col_idx).column_letter}{row_idx}"
                    ws.add_image(img, cell_address)
                    ws.cell(row=row_idx, column=img_col_idx).value = None
                except Exception as e:
                    print(f"⚠ Error inserting image in row {row_idx}: {e}")

    wb.save(excel_file)
    print(f"✅ Images embedded into {excel_file}")

def scrape_accommodation(property_type):
    filter_code = ACCOMMODATION_FILTERS.get(property_type.lower())
    if not filter_code:
        print(f"❌ '{property_type}' is not a supported accommodation type.")
        return

    max_pages = None
    stay_length = 2
    results = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        start_date = date.today() + timedelta(days=1)
        end_date = date.today() + timedelta(days=365)

        current_date = start_date
        while current_date <= end_date:
            checkin_date = current_date.strftime("%Y-%m-%d")
            checkout_date = (current_date + timedelta(days=stay_length)).strftime("%Y-%m-%d")

            print(f"🔍 Searching {property_type}s for {checkin_date} → {checkout_date}")

            page_url = (
                f"https://www.booking.com/searchresults.en-gb.html?"
                f"checkin={checkin_date}&checkout={checkout_date}&selected_currency=USD"
                f"&ss=Pakistan&group_adults=1&no_rooms=1&group_children=0&nflt={filter_code}"
            )
            page.goto(page_url, timeout=60000)

            page_number = 1
            while max_pages is None or page_number <= max_pages:
                print(f"📄 Scraping page {page_number} for {checkin_date}")

                try:
                    page.wait_for_selector('div[data-testid="property-card"]', timeout=15000)
                except:
                    print("⚠ No properties found for this date.")
                    break

                hotels = page.locator('div[data-testid="property-card"]')
                count = hotels.count()

                for i in range(count):
                    card = hotels.nth(i)

                    hotel_name = card.locator('div[data-testid="title"]').inner_text() if card.locator('div[data-testid="title"]').count() else None
                    price = card.locator('span[data-testid="price-and-discounted-price"]').inner_text() if card.locator('span[data-testid="price-and-discounted-price"]').count() else None
                    score = card.locator('div[data-testid="review-score"] > div:nth-child(1)').inner_text() if card.locator('div[data-testid="review-score"] > div:nth-child(1)').count() else None
                    avg_review = card.locator('div[data-testid="review-score"] > div:nth-child(2) > div:nth-child(1)').inner_text() if card.locator('div[data-testid="review-score"] > div:nth-child(2) > div:nth-child(1)').count() else None
                    reviews_count = card.locator('div[data-testid="review-score"] > div:nth-child(2) > div:nth-child(2)').inner_text().split()[0] if card.locator('div[data-testid="review-score"] > div:nth-child(2) > div:nth-child(2)').count() else None
                    location = card.locator('span[data-testid="address"]').inner_text() if card.locator('span[data-testid="address"]').count() else None
                    image_url = card.locator('img').get_attribute("src") if card.locator('img').count() else None

                    image_path = None
                    if image_url:
                        safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', hotel_name if hotel_name else "hotel")
                        image_filename = f"{safe_name}_{checkin_date}.jpg"
                        image_path = download_image(image_url, f"images_{property_type}", image_filename)

                    results.append({
                        "Property Type": property_type,
                        "Check-in": checkin_date,
                        "Check-out": checkout_date,
                        "Name": hotel_name,
                        "Price": price,
                        "Score": score,
                        "Average Review": avg_review,
                        "Reviews Count": reviews_count,
                        "Location": location,
                        "Image URL": image_url,
                        "Image Path": image_path
                    })

                next_button = page.locator('button[aria-label="Next page"]')
                if next_button.count() == 0 or not next_button.is_enabled():
                    break
                next_button.click()
                page.wait_for_selector('div[data-testid="property-card"]', timeout=15000)
                page_number += 1

            current_date += timedelta(days=7)

        # Save CSV
        pd.DataFrame(results).to_csv(f'pakistan_{property_type}_1year.csv', index=False)

        # Save Excel & embed images
        excel_file = f'pakistan_{property_type}_1year.xlsx'
        pd.DataFrame(results).to_excel(excel_file, index=False)
        embed_images_in_excel(excel_file)

        print(f"✅ Scraped {len(results)} {property_type} entries for 1 year with images embedded in Excel.")
        browser.close()

def open_other_services(service_name):
    SERVICE_URLS = {
        "flight": "https://www.booking.com/flights",
        "car rental": "https://www.booking.com/cars",
        "airport taxi": "https://www.booking.com/taxi",
        "activities": "https://www.booking.com/things-to-do",
        "tours": "https://www.booking.com/things-to-do",
    }
    url = SERVICE_URLS.get(service_name.lower())
    if not url:
        print(f"❌ '{service_name}' not supported for scraping.")
        return

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()
        print(f"🌐 Opening {service_name} page...")
        page.goto(url, timeout=60000)
        page.screenshot(path=f"{service_name}_page.png")
        print(f"📸 Saved screenshot of {service_name} page.")
        browser.close()

if __name__ == "__main__":
    service = input("Enter service (hotel, apartment, resort, hostel, villa, guesthouse, flight, car rental, activities): ")

    if service.lower() in ACCOMMODATION_FILTERS:
        scrape_accommodation(service)
    else:
        open_other_services(service)
